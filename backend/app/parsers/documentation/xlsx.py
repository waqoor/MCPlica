from datetime import date, datetime, time
from io import BytesIO
from uuid import UUID

from openpyxl import load_workbook

from app.core.exceptions import SourceParseError

from .common import ensure_text_limit
from .models import DocumentSection, NormalizedDocument
from .office import detect_office_document_format

_MAX_WORKBOOK_SHEETS = 100
_MAX_WORKBOOK_ROWS = 200_000
_MAX_WORKBOOK_CELLS = 2_000_000


def _value_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value).replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\\n")


def parse_xlsx(
    value: bytes,
    *,
    source_version_id: UUID,
    title: str | None,
    max_text_chars: int,
) -> NormalizedDocument:
    if detect_office_document_format(value) != "xlsx":
        raise SourceParseError("Office document is not an XLSX workbook")
    try:
        workbook = load_workbook(
            filename=BytesIO(value),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise SourceParseError("XLSX documentation is invalid or unreadable") from exc
    try:
        if len(workbook.worksheets) > _MAX_WORKBOOK_SHEETS:
            raise SourceParseError("XLSX documentation exceeds the worksheet limit")
        root = title or "Workbook"
        sections: list[DocumentSection] = []
        total_rows = 0
        total_cells = 0
        total_chars = 0
        for worksheet in workbook.worksheets:
            lines: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                total_rows += 1
                total_cells += len(row)
                if total_rows > _MAX_WORKBOOK_ROWS:
                    raise SourceParseError("XLSX documentation exceeds the row limit")
                if total_cells > _MAX_WORKBOOK_CELLS:
                    raise SourceParseError("XLSX documentation exceeds the cell limit")
                rendered = [_value_text(cell).strip() for cell in row]
                while rendered and not rendered[-1]:
                    rendered.pop()
                if not rendered:
                    continue
                line = "\t".join(rendered)
                total_chars += len(line) + 1
                if total_chars > max_text_chars:
                    raise SourceParseError("XLSX extracted text exceeds the configured limit")
                lines.append(line)
            if lines:
                sheet_text = "\n".join(lines)
                sections.append(
                    DocumentSection(
                        path=[root, worksheet.title],
                        heading=worksheet.title,
                        text=sheet_text,
                        ordinal=len(sections),
                    )
                )
        if not sections:
            raise SourceParseError("XLSX documentation contains no extractable text")
        text = "\n\n".join(f"{section.heading}\n{section.text}" for section in sections)
        ensure_text_limit(text, label="XLSX", max_text_chars=max_text_chars)
        return NormalizedDocument(
            source_version_id=source_version_id,
            title=title,
            text=text,
            sections=sections,
            metadata={
                "format": "xlsx",
                "sheet_count": len(workbook.worksheets),
                "row_count": total_rows,
            },
        )
    finally:
        workbook.close()
